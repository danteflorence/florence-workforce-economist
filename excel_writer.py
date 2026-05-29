"""
Excel workbook writer per Florence Workforce Restoration Economics v2 §8.

Produces 10-tab workbooks for a single hospital or a whole health system:
  1. Executive_Summary
  2. Inputs
  3. Pricing_Mode
  4. FICA_Offset_Model
  5. Hospital_Detail
  6. Sensitivity
  7. Market_Benchmarks
  8. Facility_Source_Data
  9. Sources_Caveats
 10. GPT_Run_Log

Public API:
    write_hospital_workbook(ccn, output_path, cal=None, cohort=None) -> Path
    write_system_workbook(health_system_id, output_path, cal=None, cohort=None) -> Path
"""

from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from pricing_engine import (
    Calibration, CohortMix, PricingMode, REQUIRED_COMPLIANCE_SENTENCE, price,
)
from pricing_batch import load_universe, price_batch, row_to_profile


# ---------------------------------------------------------------------------
# Visual style helpers
# ---------------------------------------------------------------------------

NAVY = "0B2545"
TEAL = "1E6091"
ACCENT = "2DB8A3"
LIGHT_BG = "F0F5FA"
WARN_BG = "FFF8E6"
MUTED = "6B7280"
BORDER = "E2E6EE"

thin = Side(border_style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(cell, fill_color: str = NAVY, font_color: str = "FFFFFF") -> None:
    cell.font = Font(name="Calibri", size=12, bold=True, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = box


def _label_style(cell) -> None:
    cell.font = Font(name="Calibri", size=11, color=MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _value_style(cell, bold: bool = False, fmt: Optional[str] = None) -> None:
    cell.font = Font(name="Calibri", size=11, bold=bold, color="1A2230")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        cell.number_format = fmt


def _kpi_style(cell) -> None:
    cell.font = Font(name="Calibri", size=22, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _title_style(cell) -> None:
    cell.font = Font(name="Calibri", size=18, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Per-hospital + per-system pricing aggregation
# ---------------------------------------------------------------------------

def _gather_data(
    hospital_rows: pd.DataFrame, cal: Calibration, cohort: CohortMix,
):
    priced = price_batch(hospital_rows, cohort, cal)
    feas = priced[priced["feasible"]]
    return priced, feas


def _summary_metrics(priced: pd.DataFrame, cal: Calibration) -> dict:
    feas = priced[priced["feasible"]]
    return {
        "n_hospitals": len(priced),
        "n_quotable": len(feas),
        "n_manual_review": int(priced["manual_review_flag"].sum()),
        "rn_need_total": float(feas["rn_need"].sum()),
        "median_monthly_fee": float(feas["florence_monthly_fee_per_rn"].median()) if len(feas) else 0,
        "median_fica_savings": float(feas["employer_fica_savings_per_rn_per_month"].median()) if len(feas) else 0,
        "median_effective_cost": float(feas["fica_adjusted_effective_cost_per_rn_month"].median()) if len(feas) else 0,
        "median_offset_pct": float(feas["actual_fica_offset_pct"].median()) if len(feas) else 0,
        "median_net_savings": float(feas["net_monthly_savings_per_rn"].median()) if len(feas) else 0,
        "total_monthly_fee": float(feas["monthly_florence_fee_account"].sum()),
        "total_monthly_fica": float(feas["monthly_fica_offset_account"].sum()),
        "total_monthly_effective": float(feas["monthly_effective_cost_account"].sum()),
        "total_monthly_net_savings": float(feas["monthly_net_savings_account"].sum()),
        "total_monthly_agency_avoided": float(feas["monthly_agency_avoided_account"].sum()),
        "term_florence_fee": float(feas["term_florence_fee_account"].sum()),
        "term_fica_offset": float(feas["term_fica_offset_account"].sum()),
        "term_net_savings": float(feas["term_net_savings_account"].sum()),
        "term_months": cal.term_months,
    }


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def _tab_executive_summary(wb, target_name: str, target_type: str, scope_desc: str,
                           cal: Calibration, cohort: CohortMix, m: dict) -> None:
    ws = wb.active
    ws.title = "1_Executive_Summary"
    _set_col_widths(ws, {"A": 36, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22})

    # Header band
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Florence Workforce Restoration — {target_type}: {target_name}"
    _title_style(ws["A1"])
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"Scope: {scope_desc}  ·  "
        f"Pricing mode: {cal.pricing_mode.value}  ·  "
        f"Term: {cal.term_months} months  ·  "
        f"Report date: {date.today().isoformat()}  ·  "
        f"Calibration: {cal.version}"
    )
    ws["A2"].font = Font(size=10, color=MUTED)
    ws["A2"].alignment = Alignment(horizontal="left")

    # Three-number KPI block (rows 4-7)
    ws["A4"] = "PRIMARY BUYER-FACING NUMBERS"
    _header_style(ws["A4"], NAVY)
    ws.merge_cells("A4:F4")

    kpi_labels = [
        ("①  Florence Monthly Fee / RN",  m["median_monthly_fee"],  '"$"#,##0'),
        ("②  Employer FICA Savings / RN/mo", m["median_fica_savings"], '"$"#,##0'),
        ("③  FICA-Adjusted Effective Cost / RN/mo", m["median_effective_cost"], '"$"#,##0'),
        ("④  Actual FICA Offset %", m["median_offset_pct"], "0.0%"),
        ("⑤  Net Monthly Savings / RN", m["median_net_savings"], '"$"#,##0'),
    ]
    for i, (label, value, fmt) in enumerate(kpi_labels):
        row = 5 + i
        ws[f"A{row}"] = label
        _label_style(ws[f"A{row}"])
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = value
        _value_style(ws[f"B{row}"], bold=True, fmt=fmt)
        ws[f"B{row}"].font = Font(size=14, bold=True, color=NAVY)

    # Offset achievement
    constrained = "Exact target" if abs(m["median_offset_pct"] - cal.target_offset_pct) < 0.01 else (
        "Above target (ceiling clamp)" if m["median_offset_pct"] > cal.target_offset_pct
        else "Below target (floor clamp)"
    )
    ws["A11"] = "Target offset achievement"
    _label_style(ws["A11"])
    ws.merge_cells("B11:F11")
    ws["B11"] = f"Target {cal.target_offset_pct:.0%}  |  Actual median {m['median_offset_pct']:.1%}  |  {constrained}"
    ws["B11"].font = Font(size=11, italic=True, color=NAVY)

    # Need block
    ws["A13"] = "RN NEED & TERM"
    _header_style(ws["A13"], TEAL)
    ws.merge_cells("A13:F13")
    need_rows = [
        ("Hospitals in scope", m["n_hospitals"]),
        ("Hospitals quotable", m["n_quotable"]),
        ("Hospitals flagged for manual review", m["n_manual_review"]),
        ("Total Covered RN Need (FTE)", round(m["rn_need_total"], 0)),
        ("Monthly hours per covered RN", cal.monthly_hours_rn),
        ("Annual hours per covered RN", cal.annual_hours_rn),
        ("Contract term (months)", cal.term_months),
        ("Cohort visa-exempt share (η)", cohort.eta),
        ("FICA-eligible months per nurse", cohort.eligible_months or cal.fica_eligible_months_default),
    ]
    for i, (label, value) in enumerate(need_rows):
        r = 14 + i
        ws[f"A{r}"] = label
        _label_style(ws[f"A{r}"])
        ws[f"B{r}"] = value
        _value_style(ws[f"B{r}"], fmt="#,##0")

    # Monthly bridge
    ws["A24"] = "MONTHLY BRIDGE (per RN)"
    _header_style(ws["A24"], NAVY)
    ws.merge_cells("A24:F24")
    bridge = [
        ("Florence Monthly Fee (gross)", m["median_monthly_fee"]),
        ("− Employer FICA Offset", -m["median_fica_savings"]),
        ("= FICA-Adjusted Effective Cost", m["median_effective_cost"]),
        ("+ Agency Premium Avoided", m["median_monthly_fee"] and (m["total_monthly_agency_avoided"]/max(m["rn_need_total"],1) if m["rn_need_total"] else 0)),
        ("= Net Monthly Savings per RN", m["median_net_savings"]),
    ]
    # Compute agency avoided per RN for the bridge
    agency_per_rn = m["total_monthly_agency_avoided"] / m["rn_need_total"] if m["rn_need_total"] else 0
    bridge[3] = ("+ Agency Premium Avoided", agency_per_rn)
    for i, (label, value) in enumerate(bridge):
        r = 25 + i
        ws[f"A{r}"] = label
        _label_style(ws[f"A{r}"])
        ws[f"B{r}"] = value
        _value_style(ws[f"B{r}"], bold=(i==4), fmt='"$"#,##0.00')

    # Portfolio block
    ws["A32"] = f"PORTFOLIO TOTALS ({m['n_quotable']} quotable hospitals)"
    _header_style(ws["A32"], NAVY)
    ws.merge_cells("A32:F32")
    portfolio = [
        ("Total monthly Florence billings", m["total_monthly_fee"]),
        ("Total monthly FICA offset to hospitals", m["total_monthly_fica"]),
        ("Total monthly FICA-adjusted effective cost", m["total_monthly_effective"]),
        ("Total monthly agency premium avoided", m["total_monthly_agency_avoided"]),
        ("Total monthly net hospital savings", m["total_monthly_net_savings"]),
        (f"Term ({cal.term_months}mo) Florence fee", m["term_florence_fee"]),
        (f"Term ({cal.term_months}mo) Hospital net savings", m["term_net_savings"]),
        ("Savings : Fee ratio", m["term_net_savings"] / m["term_florence_fee"] if m["term_florence_fee"] else 0),
    ]
    for i, (label, value) in enumerate(portfolio):
        r = 33 + i
        ws[f"A{r}"] = label
        _label_style(ws[f"A{r}"])
        ws.merge_cells(f"B{r}:C{r}")
        ws[f"B{r}"] = value
        fmt = '"$"#,##0' if "ratio" not in label.lower() else "0.0\"x\""
        _value_style(ws[f"B{r}"], bold=True, fmt=fmt)

    # Caveat banner (compliance sentence)
    ws["A43"] = "REQUIRED COMPLIANCE STATEMENT"
    _header_style(ws["A43"], "C97A3B", "FFFFFF")
    ws.merge_cells("A43:F43")
    ws.merge_cells("A44:F47")
    ws["A44"] = REQUIRED_COMPLIANCE_SENTENCE
    ws["A44"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A44"].font = Font(size=10, color="6B5824", italic=True)
    ws["A44"].fill = PatternFill("solid", fgColor=WARN_BG)
    ws.row_dimensions[44].height = 70


def _tab_inputs(wb, cal: Calibration, cohort: CohortMix) -> None:
    ws = wb.create_sheet("2_Inputs")
    _set_col_widths(ws, {"A": 40, "B": 18, "C": 50})
    ws["A1"] = "Pricing inputs & assumptions (editable copy of run-time calibration)"
    _title_style(ws["A1"])
    ws.merge_cells("A1:C1")

    ws["A3"] = "Parameter"; ws["B3"] = "Value"; ws["C3"] = "Reference"
    for c in ("A3", "B3", "C3"): _header_style(ws[c], NAVY)

    rows = [
        ("Pricing mode", cal.pricing_mode.value, "v2 §6"),
        ("Target FICA offset %", cal.target_offset_pct, "v2 §6.2 (default 50%)"),
        ("Price floor ($/RN/month)", cal.price_floor_monthly, "v2 §6.3 — lowered to $750 covers 99% of US"),
        ("Price ceiling ($/RN/month)", cal.price_ceiling_monthly, "v2 §6.3 (default $2,000)"),
        ("Standard fee ($/RN/month)", cal.standard_monthly_fee, "v2 §6.1 (default $1,750)"),
        ("Term (months)", cal.term_months, "v2 §4 (default 24)"),
        ("Annual hours per covered RN", cal.annual_hours_rn, "v2 §4 (1,872)"),
        ("Monthly hours per covered RN", cal.monthly_hours_rn, "v2 §4 (156)"),
        ("RN share of contracted labor", cal.rn_share_of_contracted_labor, "v2 §5.3 (default 80%)"),
        ("Coverage / Displacement target", cal.coverage_fill_factor, "v2 §5.3 (default 90%)"),
        ("Agency displacement factor", cal.agency_displacement_factor, "v2 §5.4 (default 100%)"),
        ("Cohort visa-exempt share η", cohort.eta, "F-1 default: 1.0 per IRC §3121(b)(19)"),
        ("FICA-eligible months (per nurse)", cohort.eligible_months or cal.fica_eligible_months_default, "F-1 NRA period (up to ~5 calendar years)"),
        ("Immigration add-on enabled", cal.immigration_addon_enabled, "v2 §6.5 ($5K / 24 mo)"),
        ("Immigration add-on monthly", cal.immigration_addon_monthly, "$208.33/RN/mo when enabled"),
        ("Direct partner share", cal.direct_partner_share, "v2 — 0% direct enterprise"),
        ("AMN partner share", cal.amn_partner_share, "v2 — 20% via AMN wholesale"),
        ("SS wage base", cal.ss_wage_base, "IRS Rev. Proc. annual"),
        ("Calibration version", cal.version, "Engine semantic version"),
    ]
    for i, (label, val, ref) in enumerate(rows):
        r = 4 + i
        ws[f"A{r}"] = label
        _label_style(ws[f"A{r}"])
        ws[f"B{r}"] = val
        _value_style(ws[f"B{r}"])
        ws[f"C{r}"] = ref
        ws[f"C{r}"].font = Font(size=10, color=MUTED, italic=True)


def _tab_pricing_mode(wb, cal: Calibration) -> None:
    ws = wb.create_sheet("3_Pricing_Mode")
    _set_col_widths(ws, {"A": 35, "B": 50})
    ws["A1"] = f"Pricing mode in effect: {cal.pricing_mode.value}"
    _title_style(ws["A1"])
    ws.merge_cells("A1:B1")

    explanations = {
        PricingMode.FICA_OFFSET_TARGET: (
            "Suggested fee per RN per month = Employer FICA Savings ÷ target_offset_pct. "
            f"With target = {cal.target_offset_pct:.0%}, the suggested fee equals "
            f"{1/cal.target_offset_pct:.1f}× the employer FICA savings. "
            f"Final fee clamped to ${cal.price_floor_monthly:,.0f} ≤ fee ≤ ${cal.price_ceiling_monthly:,.0f}. "
            "When the cohort is FICA-eligible (F-1 default), this anchors Florence's price to a verifiable "
            "payroll-tax saving on the customer's books."
        ),
        PricingMode.STANDARD_FEE: (
            f"Flat fee per RN per month = ${cal.standard_monthly_fee:,.0f}, regardless of agency premium "
            "or FICA savings. Useful for low-data accounts or when a simple market price is preferred."
        ),
        PricingMode.BOUNDED_TARGET: (
            "Same as FICA_OFFSET_TARGET but with custom floor/ceiling overrides for specific accounts."
        ),
        PricingMode.MANUAL_EXCEPTION: (
            "Leadership-approved fee override. Bypasses the FICA-target calculation; "
            "manual_fee_override_monthly is used."
        ),
        PricingMode.LEGACY_V1: (
            "Product plan v1.0 (capture-rate formula): premium per hour = capture_rate × agency premium, "
            "clamped. Retained for sensitivity comparison only."
        ),
    }
    ws["A3"] = "Explanation"
    _header_style(ws["A3"], TEAL)
    ws.merge_cells("A3:B3")
    ws.merge_cells("A4:B7")
    ws["A4"] = explanations.get(cal.pricing_mode, "—")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws["A4"].font = Font(size=11)
    ws.row_dimensions[4].height = 80


def _tab_fica_model(wb, hospitals_df: pd.DataFrame, cal: Calibration) -> None:
    ws = wb.create_sheet("4_FICA_Offset_Model")
    _set_col_widths(ws, {"A": 30, "B": 18, "C": 50})
    ws["A1"] = "Employer FICA Offset Model (per RN)"
    _title_style(ws["A1"])
    ws.merge_cells("A1:C1")

    feas = hospitals_df[hospitals_df["feasible"]]
    median_wage = feas["loaded_staff_cost_per_hr"].median() if len(feas) else 0
    annual_wages = median_wage * cal.annual_hours_rn  # rough — uses loaded as proxy for visualization

    ws["A3"] = "Parameter"; ws["B3"] = "Value"; ws["C3"] = "Source / Note"
    for c in ("A3", "B3", "C3"): _header_style(ws[c], NAVY)

    rows = [
        ("Eligible annual wages (median)", annual_wages, "Hourly wage × 1,872 (v2 §5.6 hourly basis)"),
        ("Social Security wage base", cal.ss_wage_base, "IRS 2026 Rev. Proc."),
        ("Taxable SS wages (cap-applied)", min(annual_wages, cal.ss_wage_base), "= MIN(wages, SS base)"),
        ("Employer SS tax rate", 0.062, "IRS — 6.2%"),
        ("Employer Medicare rate", 0.0145, "IRS — 1.45%"),
        ("Employer SS savings (annual)", min(annual_wages, cal.ss_wage_base) * 0.062, "Per F-1 exempt RN"),
        ("Employer Medicare savings (annual)", annual_wages * 0.0145, "Per F-1 exempt RN"),
        ("Annual employer FICA savings", min(annual_wages, cal.ss_wage_base) * 0.062 + annual_wages * 0.0145, "Sum of above"),
        ("Run-rate monthly savings", (min(annual_wages, cal.ss_wage_base) * 0.062 + annual_wages * 0.0145) / 12, "÷ 12"),
        ("Cohort visa-exempt share η", 1.0, "F-1 confirmed default"),
        ("FICA-eligible months", cal.fica_eligible_months_default, "Per IRS NRA window"),
        ("IRS source", "IRC §3121(b)(19) + IRS Pub 519", "https://www.irs.gov/individuals/international-taxpayers/foreign-student-liability-for-social-security-and-medicare-taxes"),
    ]
    for i, (label, val, ref) in enumerate(rows):
        r = 4 + i
        ws[f"A{r}"] = label
        _label_style(ws[f"A{r}"])
        ws[f"B{r}"] = val
        if isinstance(val, float):
            _value_style(ws[f"B{r}"], fmt='"$"#,##0.00' if val > 1 else "0.00%")
        else:
            _value_style(ws[f"B{r}"])
        ws[f"C{r}"] = ref
        ws[f"C{r}"].font = Font(size=10, color=MUTED, italic=True)


def _tab_hospital_detail(wb, priced: pd.DataFrame) -> None:
    ws = wb.create_sheet("5_Hospital_Detail")
    cols = [
        ("ccn", "CCN", 12),
        ("name", "Hospital", 36),
        ("city", "City", 18),
        ("state", "State", 6),
        ("health_system", "System", 26),
        ("rn_need", "RN Need (FTE)", 12),
        ("loaded_staff_cost_per_hr", "Loaded Staff $/hr", 14),
        ("all_in_agency_per_hr", "Agency $/hr", 12),
        ("agency_premium_per_hr", "Agency Premium $/hr", 16),
        ("florence_monthly_fee_per_rn", "Florence $/RN/mo", 16),
        ("employer_fica_savings_per_rn_per_month", "FICA Savings $/RN/mo", 18),
        ("fica_adjusted_effective_cost_per_rn_month", "FICA-Adj Effective $/RN/mo", 22),
        ("actual_fica_offset_pct", "Actual Offset %", 12),
        ("monthly_agency_premium_avoided_per_rn", "Agency Avoided $/RN/mo", 18),
        ("net_monthly_savings_per_rn", "Net Mo Savings $/RN", 18),
        ("monthly_florence_fee_account", "Total Monthly Fee", 16),
        ("monthly_fica_offset_account", "Total Monthly FICA", 16),
        ("monthly_net_savings_account", "Total Monthly Net Savings", 22),
        ("term_florence_fee_account", "Term Florence Fee", 16),
        ("term_net_savings_account", "Term Net Savings", 16),
        ("final_fee_constrained_by", "Constrained By", 18),
        ("confidence", "Confidence", 10),
        ("data_source", "Data Source", 24),
        ("manual_review_flag", "Manual Review", 12),
    ]
    # Header row
    for i, (_, header, w) in enumerate(cols):
        c = ws.cell(row=1, column=i+1, value=header)
        _header_style(c, NAVY)
        ws.column_dimensions[get_column_letter(i+1)].width = w

    # Data rows
    for r, (_, h) in enumerate(priced.iterrows(), start=2):
        for i, (col, _, _) in enumerate(cols):
            val = h.get(col)
            if pd.isna(val): val = ""
            cell = ws.cell(row=r, column=i+1, value=val)
            if isinstance(val, float):
                if "pct" in col or "offset" in col and "monthly" not in col:
                    cell.number_format = "0.0%"
                elif "_per_hr" in col:
                    cell.number_format = '"$"#,##0.00'
                elif "$/RN" in cols[i][1] or "monthly" in col.lower() or "term" in col.lower():
                    cell.number_format = '"$"#,##0'
                else:
                    cell.number_format = "#,##0.0"
    ws.freeze_panes = "B2"


def _tab_sensitivity(wb, hospital_rows: pd.DataFrame, base_cal: Calibration) -> None:
    ws = wb.create_sheet("6_Sensitivity")
    _set_col_widths(ws, {"A": 28, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18})
    ws["A1"] = "Sensitivity Analysis (median fee / RN per month at each parameter)"
    _title_style(ws["A1"])
    ws.merge_cells("A1:F1")

    cohort = CohortMix(eta=1.0)
    targets = [0.30, 0.40, 0.50, 0.60, 0.75]
    etas = [0.0, 0.5, 1.0]

    ws["A3"] = "Target FICA Offset %"
    _header_style(ws["A3"], NAVY)
    for i, e in enumerate(etas):
        c = ws.cell(row=3, column=i+2, value=f"η = {e:.1f}")
        _header_style(c, NAVY)

    for j, t in enumerate(targets):
        ws.cell(row=4+j, column=1, value=f"{t:.0%}")
        for i, e in enumerate(etas):
            cal = Calibration(
                pricing_mode=base_cal.pricing_mode,
                target_offset_pct=t,
                price_floor_monthly=base_cal.price_floor_monthly,
                price_ceiling_monthly=base_cal.price_ceiling_monthly,
                term_months=base_cal.term_months,
                rn_share_of_contracted_labor=base_cal.rn_share_of_contracted_labor,
                coverage_fill_factor=base_cal.coverage_fill_factor,
                agency_displacement_factor=base_cal.agency_displacement_factor,
            )
            priced = price_batch(hospital_rows, CohortMix(eta=e), cal)
            feas = priced[priced["feasible"]]
            median_fee = float(feas["florence_monthly_fee_per_rn"].median()) if len(feas) else 0
            c = ws.cell(row=4+j, column=i+2, value=median_fee)
            c.number_format = '"$"#,##0'

    ws["A11"] = "Note: Sensitivity shown at median across hospitals in this workbook scope."
    ws["A11"].font = Font(size=10, color=MUTED, italic=True)


def _tab_market_benchmarks(wb, priced: pd.DataFrame) -> None:
    ws = wb.create_sheet("7_Market_Benchmarks")
    _set_col_widths(ws, {"A": 12, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 18})
    ws["A1"] = "Market Benchmarks — by state median"
    _title_style(ws["A1"])
    ws.merge_cells("A1:G1")

    grp = priced.groupby("state").agg(
        n=("ccn", "count"),
        staff=("loaded_staff_cost_per_hr", "median"),
        agency=("all_in_agency_per_hr", "median"),
        premium=("agency_premium_per_hr", "median"),
        fee=("florence_monthly_fee_per_rn", "median"),
        fica=("employer_fica_savings_per_rn_per_month", "median"),
        net=("net_monthly_savings_per_rn", "median"),
    ).reset_index().sort_values("n", ascending=False)

    headers = ["State", "Hospitals", "Median Staff $/hr", "Median Agency $/hr",
               "Median Premium $/hr", "Median Fee $/RN/mo", "Median FICA $/RN/mo"]
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=i+1, value=h)
        _header_style(c, NAVY)
    for r, (_, row) in enumerate(grp.iterrows(), start=4):
        ws.cell(row=r, column=1, value=row["state"])
        ws.cell(row=r, column=2, value=int(row["n"]))
        for i, k in enumerate(["staff", "agency", "premium", "fee", "fica"]):
            c = ws.cell(row=r, column=3+i, value=float(row[k]) if pd.notna(row[k]) else 0)
            c.number_format = '"$"#,##0.00' if i < 3 else '"$"#,##0'


def _tab_facility_source_data(wb, priced: pd.DataFrame) -> None:
    ws = wb.create_sheet("8_Facility_Source_Data")
    cols = [
        ("ccn", "CCN", 12),
        ("name", "Hospital", 36),
        ("city", "City", 18),
        ("state", "State", 6),
        ("county", "County", 18),
        ("hospital_type", "Type", 22),
        ("ownership", "Ownership", 24),
        ("hcris_total_fte", "HCRIS Total FTE", 14),
        ("contract_labor_dollars", "HCRIS Contract Labor $", 22),
        ("contract_labor_intensity", "CL Intensity %", 14),
        ("operating_margin", "Operating Margin %", 16),
        ("rn_need", "Derived RN Need", 14),
        ("loaded_staff_cost_per_hr", "Loaded Staff $/hr", 16),
        ("all_in_agency_per_hr", "All-in Agency $/hr", 16),
        ("data_source", "Wage/Agency Source", 32),
        ("confidence", "Data Confidence", 14),
    ]
    for i, (_, header, w) in enumerate(cols):
        c = ws.cell(row=1, column=i+1, value=header)
        _header_style(c, NAVY)
        ws.column_dimensions[get_column_letter(i+1)].width = w
    for r, (_, h) in enumerate(priced.iterrows(), start=2):
        for i, (col, _, _) in enumerate(cols):
            val = h.get(col)
            if pd.isna(val): val = ""
            cell = ws.cell(row=r, column=i+1, value=val)
            if isinstance(val, float):
                if "intensity" in col or "margin" in col or col == "confidence":
                    cell.number_format = "0.0%" if "pct" in col or "intensity" in col or "margin" in col else "0.00"
                else:
                    cell.number_format = '"$"#,##0'
    ws.freeze_panes = "B2"


def _tab_sources_caveats(wb) -> None:
    ws = wb.create_sheet("9_Sources_Caveats")
    _set_col_widths(ws, {"A": 30, "B": 90})
    ws["A1"] = "Sources, Methodology, and Compliance"
    _title_style(ws["A1"])
    ws.merge_cells("A1:B1")

    sections = [
        ("Methodology", "Florence Workforce Restoration Economics v2.0 (May 28, 2026). "
         "Monthly Per-Nurse ROI Model with 50% FICA-Offset Target Pricing."),
        ("Required Compliance Statement", REQUIRED_COMPLIANCE_SENTENCE),
        ("Hospital roster", "CMS Hospital General Information (data.cms.gov, refreshed quarterly)."),
        ("Hospital labor data", "CMS HCRIS Hospital Provider Cost Report 2023 — per-hospital FTE, "
         "total salaries, wage-related costs, contract labor."),
        ("RN wage benchmarks", "BLS OEWS state-level placeholder (v0.5). Production: BLS OEWS API by MSA + occupation 29-1141."),
        ("Agency rate benchmarks", "Hybrid: CommonSpirit internal anchors for matched hospitals, "
         "state-median imputation for the remainder."),
        ("FICA exemption — primary source",
         "IRS — Foreign Student Liability for Social Security and Medicare Taxes: "
         "https://www.irs.gov/individuals/international-taxpayers/foreign-student-liability-for-social-security-and-medicare-taxes"),
        ("FICA exemption — statutory", "IRC §3121(b)(19); IRS Publication 519; IRS Publication 15 (Circular E)."),
        ("Confidence tiers", "1.00 customer-disclosed · 0.85 HCRIS-derived · 0.60 state-imputed (CS anchor) · 0.40 national-imputed · <0.50 triggers STANDARD_FEE fallback."),
        ("Manual review", "Only fires for genuinely-negative agency premium. Low-confidence agency-rate rows are priced at standard fee with caveat pending customer disclosure."),
        ("Caveats", "Not legal, tax, immigration, accounting, or payroll advice. All tax and immigration assumptions should be confirmed by the customer and its advisors before use as contractual commitments."),
    ]
    for i, (label, body) in enumerate(sections):
        r = 3 + i * 2
        ws[f"A{r}"] = label
        _label_style(ws[f"A{r}"])
        ws[f"A{r}"].font = Font(bold=True, color=NAVY)
        ws.merge_cells(f"B{r}:B{r}")
        ws[f"B{r}"] = body
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 36


def _tab_run_log(wb, target_name: str, scope_desc: str, cal: Calibration, cohort: CohortMix) -> None:
    ws = wb.create_sheet("10_Run_Log")
    _set_col_widths(ws, {"A": 30, "B": 60})
    ws["A1"] = "Run Log (audit trail)"
    _title_style(ws["A1"])
    ws.merge_cells("A1:B1")

    rows = [
        ("Target", target_name),
        ("Scope", scope_desc),
        ("Generated at", datetime.now().isoformat(timespec="seconds")),
        ("Engine version", cal.version),
        ("Pricing mode", cal.pricing_mode.value),
        ("Cohort η", cohort.eta),
        ("FICA eligible months", cohort.eligible_months or cal.fica_eligible_months_default),
        ("Term months", cal.term_months),
        ("Target offset %", f"{cal.target_offset_pct:.0%}"),
        ("Floor / Ceiling", f"${cal.price_floor_monthly:,.0f} – ${cal.price_ceiling_monthly:,.0f}"),
        ("Standard fee", f"${cal.standard_monthly_fee:,.0f}"),
        ("Low-confidence fallback to standard", cal.use_standard_fee_for_low_confidence),
        ("Immigration add-on enabled", cal.immigration_addon_enabled),
        ("Direct partner share", f"{cal.direct_partner_share:.0%}"),
        ("AMN partner share", f"{cal.amn_partner_share:.0%}"),
    ]
    for i, (label, val) in enumerate(rows):
        r = 3 + i
        ws[f"A{r}"] = label; _label_style(ws[f"A{r}"])
        ws[f"B{r}"] = str(val); _value_style(ws[f"B{r}"])
        ws[f"B{r}"].alignment = Alignment(horizontal="left")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_hospital_workbook(
    ccn: str, output_path: Path,
    cal: Optional[Calibration] = None, cohort: Optional[CohortMix] = None,
) -> Path:
    cal = cal or Calibration()
    cohort = cohort or CohortMix(eta=1.0)
    u = load_universe()
    u["ccn"] = u["ccn"].astype(str).str.zfill(6)
    ccn = str(ccn).zfill(6)
    h_rows = u[u["ccn"] == ccn]
    if h_rows.empty:
        raise ValueError(f"CCN {ccn} not found in universe")
    priced, _ = _gather_data(h_rows, cal, cohort)
    m = _summary_metrics(priced, cal)

    target_name = h_rows.iloc[0]["name"]
    scope = f"{h_rows.iloc[0]['city']}, {h_rows.iloc[0]['state']} · CCN {ccn}"

    wb = Workbook()
    _tab_executive_summary(wb, target_name, "Hospital", scope, cal, cohort, m)
    _tab_inputs(wb, cal, cohort)
    _tab_pricing_mode(wb, cal)
    _tab_fica_model(wb, priced, cal)
    _tab_hospital_detail(wb, priced)
    _tab_sensitivity(wb, h_rows, cal)
    _tab_market_benchmarks(wb, priced)
    _tab_facility_source_data(wb, priced)
    _tab_sources_caveats(wb)
    _tab_run_log(wb, target_name, scope, cal, cohort)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_system_workbook(
    health_system_id: str, output_path: Path,
    cal: Optional[Calibration] = None, cohort: Optional[CohortMix] = None,
) -> Path:
    cal = cal or Calibration()
    cohort = cohort or CohortMix(eta=1.0)
    u = load_universe()
    sys_rows = u[u["health_system_id"] == health_system_id]
    if sys_rows.empty:
        # Try by canonical name as a fallback
        sys_rows = u[u["health_system"] == health_system_id]
        if sys_rows.empty:
            raise ValueError(f"Health system {health_system_id!r} not found")
    priced, _ = _gather_data(sys_rows, cal, cohort)
    m = _summary_metrics(priced, cal)

    target_name = sys_rows.iloc[0]["health_system"]
    states = sorted(sys_rows["state"].unique())
    scope = (f"{len(sys_rows)} hospitals in {len(states)} states "
             f"({', '.join(states[:5])}{'...' if len(states) > 5 else ''})")

    wb = Workbook()
    _tab_executive_summary(wb, target_name, "Health System", scope, cal, cohort, m)
    _tab_inputs(wb, cal, cohort)
    _tab_pricing_mode(wb, cal)
    _tab_fica_model(wb, priced, cal)
    _tab_hospital_detail(wb, priced)
    _tab_sensitivity(wb, sys_rows, cal)
    _tab_market_benchmarks(wb, priced)
    _tab_facility_source_data(wb, priced)
    _tab_sources_caveats(wb)
    _tab_run_log(wb, target_name, scope, cal, cohort)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # Smoke test: generate Kaiser system + one single-hospital workbook
    out = Path("proposals")
    print("Generating Kaiser system workbook...")
    p1 = write_system_workbook("kaiser_permanente", out / "Kaiser_Permanente_v2.xlsx")
    print(f"  Wrote {p1}")

    print("Generating Kaiser Sunnyside (OR) single-hospital workbook...")
    u = load_universe()
    sunnyside = u[u["name"].str.contains("SUNNYSIDE", na=False)]
    if len(sunnyside):
        p2 = write_hospital_workbook(sunnyside.iloc[0]["ccn"], out / "Kaiser_Sunnyside.xlsx")
        print(f"  Wrote {p2}")
